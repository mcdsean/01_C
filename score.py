#
# Running this script will automatically score the KDM and Juliet test case suites
#
# 2016-12-01 smcdonagh@keywcorp.com: initial version
#
import os, re, argparse, shutil, py_common, operator
import xml.etree.ElementTree as elemTree

from hashlib import sha1
from time import strftime
from suite import Suite, TestCase
from operator import itemgetter
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, PatternFill, Font, Alignment
from openpyxl.chart import BarChart, LineChart
from openpyxl.chart import PieChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.chart.text import RichText
from openpyxl.drawing.text import RichTextProperties, Paragraph, ParagraphProperties, CharacterProperties

# Global for command line argument
normalize_juliet_false_scoring = False

TOOL_NAME = 'fortify'
XML_OUTPUT_DIR = 'xmls'
WID_DELIMITER_FORTIFY = ':'


def format_workbook():
    # column titles
    hit_sheet_titles = ['CWE', 'Type', 'T/F', 'File Name', 'Line #', 'Function', 'SCORE', 'Opps', '%', 'Opportunities']
    for idx, title in enumerate(hit_sheet_titles):
        set_appearance(ws3, 1, idx + 1, 'fg_fill', 'C9C9C9')
        ws3.cell(row=1, column=idx + 1).value = title
        ws3.cell(row=1, column=idx + 1).alignment = Alignment(horizontal='center')
    # column widths
    ws1_col_widths = [('A', 8), ('B', 8), ('C', 8), ('D', 8), ('E', 8), ('F', 8), ('G', 8), ('H', 22), ('I', 8)]
    ws2_col_widths = [('A', 8), ('B', 6), ('C', 6), ('D', 5), ('E', 6), ('F', 6), ('G', 38), ('H', 62), ('I', 95)]
    ws3_col_widths = [('A', 8), ('B', 6), ('C', 6), ('D', 108), ('E', 6), ('F', 17), ('G', 6), ('H', 6), ('I', 8),
                      ('J', 12), ('K', 12), ('L', 12), ('M', 12)]
    ws4_col_widths = [('A', 18), ('B', 6), ('C', 6), ('D', 6), ('E', 6), ('F', 10), ('G', 6), ('H', 6), ('I', 6)]
    ws5_col_widths = [('A', 8), ('B', 8), ('C', 8), ('D', 8), ('E', 8), ('F', 8), ('G', 8), ('H', 7), ('I', 7),
                      ('J', 5), ('K', 7), ('L', 7), ('M', 5), ('AC', 12)]
    # column widths
    for ws1_col_id, ws1_col_width in ws1_col_widths:
        ws1.column_dimensions[ws1_col_id].width = ws1_col_width
    for ws2_col_id, ws2_col_width in ws2_col_widths:
        ws2.column_dimensions[ws2_col_id].width = ws2_col_width
    for ws3_col_id, ws3_col_width in ws3_col_widths:
        ws3.column_dimensions[ws3_col_id].width = ws3_col_width
    for ws4_col_id, ws4_col_width in ws4_col_widths:
        ws4.column_dimensions[ws4_col_id].width = ws4_col_width
    for ws5_col_id, ws5_col_width in ws5_col_widths:
        ws5.column_dimensions[ws5_col_id].width = ws5_col_width
    # zoom
    ws1.sheet_view.zoomScale = 80
    ws2.sheet_view.zoomScale = 80
    ws3.sheet_view.zoomScale = 70
    ws4.sheet_view.zoomScale = 80
    ws5.sheet_view.zoomScale = 70
    # gridlines
    ws1.sheet_view.showGridLines = False
    ws2.sheet_view.showGridLines = False
    ws3.sheet_view.showGridLines = False
    ws4.sheet_view.showGridLines = False
    ws5.sheet_view.showGridLines = False
    # freeze panes
    ws2.freeze_panes = ws2['B2']
    ws3.freeze_panes = ws3['A2']
    ws4.freeze_panes = ws4['A2']
    ws5.freeze_panes = ws5['N2']
    # merge 'Opportunities' title cells
    ws3.merge_cells('J1:M1')
    # hide helper columns
    for col in ['AC', 'AD']:
        ws1.column_dimensions[col].hidden = True


def get_schemas(suite_dat):
    schemas = {}
    weakness_id_schemas = []

    tag_ids = getattr(suite_dat, 'tag_info')

    # get xml schemas from vendor input file
    for idx, content in enumerate(tag_ids):
        schema = 'ns1:' + getattr(suite_dat, 'tag_info')[idx][1].replace('/', '/ns1:')

        # 'Item' & 'Tag or Attribute'
        item = content[0].lower()
        tag_or_attribute = content[2].lower()

        # 'Finding_Type'
        if item == 'finding_type':
            if tag_or_attribute == 'tag':
                schemas['finding_type_schema'] = schema
            continue

        # 'File_Name', 'Line_Number' & 'Function_Name'
        if item == 'file_name' or item == 'line_number' or item == 'function_name':
            if tag_or_attribute == 'tag':
                schemas[item + '_schema'] = schema
            elif tag_or_attribute == 'attribute':
                schemas[item + '_schema'] = schema.rsplit('/', 1)[0]
                schemas[item + '_attrib'] = schema.rsplit(':', 1)[1]
            continue

        # 'Weakness_ID_'s
        if 'weakness' in item:
            weakness_id_schemas.append('ns1:' + str(tag_ids[idx][1]).replace('/', '/ns1:'))

    return schemas, weakness_id_schemas


def score_xmls(suite_dat):
    ns = {}
    wid_pieces_that_hit = []

    schemas, weakness_id_schemas = get_schemas(suite_dat)

    for xml_project in suite_data.xml_projects:
        used_wids = []
        test_cases = []
        test_case_files = []
        file_paths = []

        test_case_objects = []

        # read namespace from the first xml since it will be the same for all other xmls
        xml_path = os.path.join(os.getcwd(), 'xmls', getattr(xml_project, 'new_xml_name'))
        tree = elemTree.parse(xml_path)
        root = tree.getroot()
        ns['ns1'] = root.tag.split('}')[0].replace('{', '')

        setattr(suite_dat, 'name_space', ns)  # todo: we only need this one time

        print('XML', xml_path)

        # get the acceptable wids for this xml
        good_wids = getattr(xml_project, 'acceptable_weakness_ids')
        test_case_type = getattr(xml_project, 'tc_type')
        tool_name = getattr(suite_data, 'tool_name')

        # 1. parse thru each row in this xml looking for good wids, and get the filename & line #
        # for vuln in root.findall('./' + finding_type_schema, ns):
        for vuln in root.findall('./' + schemas['finding_type_schema'], ns):
            # todo: add condition for tag vs attribute; need flags in switch statement above
            del wid_pieces_that_hit[:]

            # 2. get relative path/filename and line number for this row in this xml

            try:
                file_path = vuln.find(schemas['file_name_schema'], ns).attrib[schemas['file_name_attrib']]
            except AttributeError:
                print('Hit Has No Path:', xml_path)
                continue

            # exclude support files
            if not file_path.startswith('T/') and not file_path.startswith('F/'):
                continue
            line_number = vuln.find(schemas['line_number_schema'], ns).attrib[schemas['line_number_attrib']]
            function_name = vuln.find(schemas['function_name_schema'], ns).attrib[schemas['function_name_attrib']]

            # 3. get all pieces of the wid for this row in the xml
            for idx, weakness_id in enumerate(weakness_id_schemas):
                wid_piece = vuln.find(weakness_id_schemas[idx], ns)
                if wid_piece is not None:
                    wid_pieces_that_hit.append(wid_piece.text)

            # 4. look at each non-empty cell in the spreadsheet for acceptable wids
            for good_wid in good_wids:
                if tool_name == 'fortify':
                    # split each wid based on it's delimiter
                    good_wid_pieces = good_wid.split(WID_DELIMITER_FORTIFY)
                else:
                    good_wid_pieces = good_wid

                # todo: 5/5/17 optimize, move on to next upon first blank cell in row
                # 5. if the current cell in spreadsheet does not contain a cwe # or is blank, move on
                if good_wid != 'None' and not good_wid.isdigit():
                    # 6. see if ALL of the pieces for this row match this cell's good wid
                    if set(wid_pieces_that_hit) != set(good_wid_pieces):
                        continue
                    else:
                        # this is a good wid so add it to the list if it is not already there
                        if good_wid not in used_wids:
                            used_wids.append(good_wid)

                        test_case_files.append([file_path, int(line_number)])

                        if test_case_type == 'juliet':
                            if suite_language == 'c':
                                test_case_name = re.sub('[a-z]?\.\w+$', '', file_path)
                                # reduce juliet function name to 'good ...' portion
                                function_name = function_name.rpartition('_')[2]
                            elif suite_language == 'cpp':
                                test_case_name = re.sub('([a-z]?\.\w+$)|(_good.*)|(_bad.*)', '', file_path)
                                # reduce juliet function name to 'good ...' portion
                                if function_name == 'action':
                                    function_name = file_path.rpartition('_')[2][:-4]
                                else:
                                    function_name = function_name.rpartition('_')[2]

                        elif test_case_type == 'kdm':
                            # todo: 5/5/17 reduce kdm name for display in ws3 (similar to juliet above)
                            test_case_name = re.sub('[_a]?\.\w+$', '', file_path)
                        else:
                            print('ERROR: NO TEST CASE NAME FOUND')
                            test_case_name = ''

                        file_paths.append(file_path)

                        if test_case_name not in test_cases:
                            # create a new test case object
                            new_tc_obj = TestCase(test_case_name, xml_project.tc_type, xml_project.true_false,
                                                  suite_language)
                            new_tc_obj.hit_data.append([file_path, line_number, function_name])
                            test_case_objects.append(new_tc_obj)

                            # add the new test case object to the xml project list
                            setattr(xml_project, 'test_cases', test_case_objects)
                            test_cases.append(test_case_name)
                            # store the number of hits for this test case
                            name = new_tc_obj.test_case_name
                            suite_dat.suite_hit_data[name] = len(new_tc_obj.hit_data)

                        else:  # todo: 5/3/17, consider using a dictionary here or default dict for speed
                            # update existing test case object
                            for test_case_object in test_case_objects:
                                if test_case_object.test_case_name == test_case_name:
                                    print('CPP_TEST_CASE_NAME_NOT_FIRST_OCCURRANCE________', test_case_name)
                                    hit_data = getattr(test_case_object, 'hit_data')
                                    hit_data.append([file_path, line_number, function_name])
                                    # update the number of hits for this test case
                                    name = test_case_object.test_case_name
                                    suite_dat.suite_hit_data[name] = len(test_case_object.hit_data)
                                    break

                # empty acceptable wid cell on spreadsheet so move on
                else:
                    continue

        if test_case_type == 'juliet' and xml_project.true_false == 'FALSE':
            score = 0  # todo: 5/5/7 juliet/false will be calculated seperately
            # score = calculate_juliet_false_xml_score(xml_project)

        else:
            # juliet(true) and kdm counts, one hit per test case
            score = len(set(test_cases))

        # store data for each xml project
        setattr(xml_project, 'num_of_hits', score)
        setattr(xml_project, 'used_wids', used_wids)
        setattr(xml_project, 'test_case_files_that_hit', file_paths)

        print('XML_PROJECT_NAME:', xml_project.new_xml_name, 'SCORE:', score)


def calculate_test_case_score(test_case_obj):
    valid_hits = []

    for valid_hit_data in test_case_obj.hit_data:
        valid_hits.append(valid_hit_data[2])
    score = len(set(valid_hits))
    test_case_obj.score = score

    if test_case_obj.tc_type == 'juliet' and test_case_obj.true_false == 'FALSE':
        setattr(test_case_obj, 'score', score)


def calculate_test_case_percent_hits(test_case_obj):
    if test_case_obj.opp_counts == 0:
        print('HITS_WITH_NO_OPPS', test_case_obj.test_case_name)
    else:
        percent = test_case_obj.score / test_case_obj.opp_counts
        test_case_obj.percent = percent


def collect_hit_data(suite_dat):
    # file name, line number, and enclosing function
    hit_data = []

    # collect all valid hit data to be displayed
    for xml_project in suite_dat.xml_projects:
        test_case_objects = xml_project.test_cases
        print('collecting hit data for project-----', xml_project.new_xml_name)

        for test_case_obj in test_case_objects:
            # calculate the score for this test case
            # todo: 5/5/7 we only need this for juliet, false?
            # todo: 5/5/7 bug found in sheet for juliet, true, get 200% of four hits?
            calculate_test_case_score(test_case_obj)

            # calculate the percent hits for this test case
            calculate_test_case_percent_hits(test_case_obj)

            # build the columns for ws3
            for data1 in test_case_obj.hit_data:
                # A , B, C, D-F, G, H, I
                hit_data_columns = [xml_project.cwe_id_padded] + \
                                   [xml_project.tc_type] + \
                                   [xml_project.true_false] + \
                                   data1 + \
                                   [test_case_obj.score] + \
                                   [test_case_obj.opp_counts] + \
                                   [str(round(test_case_obj.percent * 100, 1)) + ' %']
                # J-M, opportunities
                hit_data_columns.extend(test_case_obj.opp_names)
                # add to composite list for writing to ws3
                hit_data.append(hit_data_columns)

            # process only juliet, false
            if xml_project.tc_type == 'juliet' and xml_project.true_false == 'FALSE':
                xml_project.num_of_hits += test_case_obj.score
                xml_project.tc_count += test_case_obj.opp_counts

        # todo: encapusalate with try/except block for divide by zero possibility (although unlikely to occur)
        # try:
        xml_project.percent_hits = str(round((xml_project.num_of_hits / xml_project.tc_count) * 100, 1)) + ' '
        # except:
        # todo: add error checking for 'DIVIDE BY ZERO ERROR'
        print('Collecting Hit Data for:', xml_project.new_xml_name)

    # sort hits by file name and then line number
    hit_data = sorted(hit_data, key=operator.itemgetter(3, 4))

    group_hit_data(suite_dat, hit_data)


def group_hit_data(suite_dat, hit_data):
    list_of_dicts = []

    suite_data.suite_hit_data_complete = hit_data
    good_data = []

    # dedupe data
    for hit_dat in suite_data.suite_hit_data_complete:
        # only juliet/false have 'good...' opportunities
        if hit_dat[1] == 'juliet' and hit_dat[2] == 'FALSE':
            # good_data.append(hit_dat[:-7])
            good_data.append(hit_dat)

    good_data_unique = remove_dups(good_data)

    for idx, data_unique in enumerate(good_data_unique):
        # name of containing function
        name = data_unique[5]
        # score
        hits = int(data_unique[6])
        # opps
        opps = int(data_unique[7])

        list_of_dicts = update_list_of_dicts(list_of_dicts, name, hits, opps)

    list_of_dicts = sorted(list_of_dicts, key=itemgetter('name'))

    b2g_idx, b2g_row_start = 0, 0
    g2b_idx, g2b_row_start = 0, 0
    g2b_hits_total, b2g_hits_total = 0, 0
    g2b_opps_total, b2g_opps_total = 0, 0

    hit_analytics_titles = ['Encapsulating Function', 'Hits', 'Misses', 'Opps', '%-Hits', 'Group', 'HITS', 'OPPS',
                            '%-grp']
    for idx, title in enumerate(hit_analytics_titles):
        set_appearance(ws4, 1, idx + 1, 'fg_fill', 'C9C9C9')
        ws4.cell(row=1, column=idx + 1).value = title
        ws4.cell(row=1, column=4).alignment = Alignment(horizontal='center')
        if idx > 4:
            ws4.cell(row=1, column=idx).alignment = Alignment(horizontal='center')
            ws4.cell(row=2, column=idx).alignment = Alignment(horizontal='center')
            ws4.cell(row=3, column=idx).alignment = Alignment(horizontal='center')

    # write to 'hit analytics' summary sheet
    for idx, hits1 in enumerate(list_of_dicts):
        if 'helperGood' in hits1['name']:
            # todo: log all of these and provide used more specifics w/ location, etc.
            suite_data.manual_review_recommendataion = ' * Manual Review Required for ' + hits1['name']

        # write summary data
        if hits1['opps'] == 0:
            print('HITS_WITH_NO_OPPS_1', hits1['name'])
            percent = 0
        else:
            percent = hits1['hits'] / hits1['opps'] * 100

        misses = hits1['opps'] - hits1['hits']
        ws4.cell(row=idx + 2, column=1).value = hits1['name']
        ws4.cell(row=idx + 2, column=2).value = hits1['hits']
        ws4.cell(row=idx + 2, column=3).value = misses
        ws4.cell(row=idx + 2, column=4).value = hits1['opps']
        ws4.cell(row=idx + 2, column=5).value = '%0.0f' % percent + '%'
        ws4.cell(row=idx + 2, column=6).value = hits1['name']

        # todo: 08312017
        # get row indexing for B2G
        if 'B2G' in hits1['name']:
            if b2g_row_start == 0:
                b2g_row_start = idx + 2

            b2g_idx += 1
            b2g_hits_total += hits1['hits']
            b2g_opps_total += hits1['opps']
            b2g_percent_total = b2g_hits_total / b2g_opps_total * 100
            ws4.cell(row=b2g_row_start, column=7).value = b2g_hits_total
            ws4.cell(row=b2g_row_start, column=8).value = b2g_opps_total
            ws4.cell(row=b2g_row_start, column=9).value = '%0.0f' % b2g_percent_total + '%'

        # get row indexing for G2B
        elif 'G2B' in hits1['name']:
            if g2b_row_start == 0:
                g2b_row_start = idx + 2

            g2b_idx += 1
            g2b_hits_total += hits1['hits']
            g2b_opps_total += hits1['opps']
            g2b_percent_total = g2b_hits_total / g2b_opps_total * 100
            ws4.cell(row=g2b_row_start, column=7).value = g2b_hits_total
            ws4.cell(row=g2b_row_start, column=8).value = g2b_opps_total
            ws4.cell(row=g2b_row_start, column=9).value = '%0.0f' % g2b_percent_total + '%'
        else:
            ws4.cell(row=idx + 2, column=7).value = hits1['hits']
            ws4.cell(row=idx + 2, column=8).value = hits1['opps']
            ws4.cell(row=idx + 2, column=9).value = '%0.0f' % percent + '%'

    # merge and align cells
    for col_idx, row in enumerate(hit_analytics_titles):
        if col_idx > 4:
            ws4.cell(row=1 + 1, column=10).value = ws4.cell(row=1, column=6).value
            ws4.cell(row=2 + 1, column=10).value = ws4.cell(row=2, column=6).value
            ws4.cell(row=3 + 1, column=10).value = ws4.cell(row=3, column=6).value
            ws4.cell(row=b2g_row_start + 1, column=10).value = ws4.cell(row=b2g_row_start, column=6).value
            ws4.cell(row=b2g_row_start + 1 + 1, column=10).value = ws4.cell(row=g2b_row_start, column=6).value
            ws4.cell(row=1 + 1, column=11).value = ws4.cell(row=1, column=8).value
            ws4.cell(row=2 + 1, column=11).value = ws4.cell(row=2, column=8).value
            ws4.cell(row=3 + 1, column=11).value = ws4.cell(row=3, column=8).value
            ws4.cell(row=b2g_row_start + 1, column=11).value = ws4.cell(row=b2g_row_start, column=8).value
            ws4.cell(row=b2g_row_start + 1 + 1, column=11).value = ws4.cell(row=g2b_row_start, column=8).value

            ws4.merge_cells(start_row=b2g_row_start, start_column=col_idx + 1, end_row=b2g_row_start + b2g_idx - 1,
                            end_column=col_idx + 1)
            ws4.merge_cells(start_row=g2b_row_start, start_column=col_idx + 1, end_row=g2b_row_start + g2b_idx - 1,
                            end_column=col_idx + 1)
            ws4.cell(row=b2g_row_start, column=col_idx + 1).alignment = Alignment(horizontal='center',
                                                                                  vertical='center')
            ws4.cell(row=g2b_row_start, column=col_idx + 1).alignment = Alignment(horizontal='center',
                                                                                  vertical='center')

    # color cells
    for idx, hits1 in enumerate(list_of_dicts):
        ws4.cell(row=idx + 2, column=4).alignment = Alignment(horizontal='right', vertical='center')
        ws4.cell(row=idx + 2, column=5).alignment = Alignment(horizontal='center', vertical='center')
        ws4.cell(row=idx + 2, column=6).alignment = Alignment(horizontal='center', vertical='center')

        for col_idx, val in enumerate(hit_analytics_titles):
            if col_idx == 1:  # hits
                set_appearance(ws4, idx + 2, col_idx + 1, 'fg_fill', 'E2EFDA')  # light green
                set_appearance(ws4, idx + 2, col_idx + 1, 'font_color', '358254')  # dark green
            elif col_idx == 2:  # misses
                set_appearance(ws4, idx + 2, col_idx + 1, 'fg_fill', 'FCE4D6')  # light red
                set_appearance(ws4, idx + 2, col_idx + 1, 'font_color', '990000')  # dark red
            else:
                if 'B2G' in hits1['name']:
                    set_appearance(ws4, idx + 2, col_idx + 1, 'fg_fill', 'F2F2F2')  # light gray
                elif 'G2B' in hits1['name']:
                    set_appearance(ws4, idx + 2, col_idx + 1, 'fg_fill', 'D0CECE')  # dark gray
                else:
                    set_appearance(ws4, idx + 2, col_idx + 1, 'fg_fill', 'FFFFFF')  # white

        for i in range(2, 9):
            ws4.cell(row=idx + 2, column=i).number_format = '#,##0'

    create_hit_charts()

    print('Writing hit data to sheet ... please stand by, thank you for your patience!')
    write_hit_data(suite_dat, hit_data)


def create_hit_charts():
    hit_bar_chart = BarChart(gapWidth=50)
    hit_bar_chart.type = 'col'
    hit_bar_chart.style = 12
    hit_bar_chart.grouping = 'stacked'
    hit_bar_chart.overlap = 100
    hit_bar_chart.title = 'Function Hits vs. Opportunities (Juliet/False Only)'
    hit_bar_chart.y_axis.title = 'Total Hits per Group'

    g2b_data = Reference(ws4, min_col=2, min_row=1, max_row=15, max_col=2)
    hit_bar_chart.add_data(g2b_data, titles_from_data=True)
    g2b_data = Reference(ws4, min_col=3, min_row=1, max_row=15, max_col=3)
    hit_bar_chart.add_data(g2b_data, titles_from_data=True)

    s5 = hit_bar_chart.series[0]
    s5.graphicalProperties.line.solidFill = '000000'
    s5.graphicalProperties.solidFill = 'C5E0B4'  # light green

    s5 = hit_bar_chart.series[1]
    s5.graphicalProperties.line.solidFill = '000000'
    s5.graphicalProperties.solidFill = 'F8CBAD'  # light orange

    cats = Reference(ws4, min_col=1, min_row=2, max_row=15)
    hit_bar_chart.set_categories(cats)
    hit_bar_chart.height = 12
    hit_bar_chart.width = 24
    ws4.add_chart(hit_bar_chart, 'J2')

    # set x-axis label orientation to 45 degrees
    hit_bar_chart.x_axis.txPr = RichText(bodyPr=RichTextProperties(
        anchor='ctr', anchorCtr='1', rot='-2700000', spcFirstLastPara='1', vertOverflow='ellipsis', wrap='square'),
        p=[Paragraph(pPr=ParagraphProperties(defRPr=CharacterProperties()), endParaRPr=CharacterProperties())])

    # Groups chart
    pie = PieChart()
    pie.height = 10
    pie.width = 13.9
    pie.title = 'Hits by Group'

    labels = Reference(ws4, min_col=10, min_row=3, max_row=6)
    data = Reference(ws4, min_col=11, min_row=2, max_row=6)
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)

    # color the slices
    pie_slice = pie.series[0]
    pt = DataPoint(idx=2)
    # todo keep for now as reference
    # pt.graphicalProperties.line.solidFill = 'FFFFFF' # white
    pt.graphicalProperties.solidFill = 'BFBFBF'  # gray
    pie_slice.dPt.append(pt)

    ws4.add_chart(pie, 'A16')


def update_list_of_dicts(list_data, name, hits, opps):
    found = False
    # update the dictionary if it already exists
    if len(list_data) > 0:
        for d in list_data:
            if d['name'] == name:
                d['hits'] += hits
                d['opps'] += opps
                found = True
                break
    # create new dictionary in the list if it does not exists
    if not found:
        list_data.append({'name': name, 'hits': hits, 'opps': opps})

    return list_data


def write_hit_data(suite_dat, hit_data):
    row = 1
    file_seen = set()
    file_name_dups = []

    # column alignments
    horizontal_left = [4]
    horizontal_right = [9]

    for hit in hit_data:
        col = 1
        for cell in hit:
            # write hit data to cells in ws3
            ws3.cell(row=row + 1, column=col).value = cell

            # set the alignment based on column
            if col in horizontal_right:
                ws3.cell(row=row + 1, column=col).alignment = Alignment(horizontal='right', vertical='center')
            elif col not in horizontal_left:
                ws3.cell(row=row + 1, column=col).alignment = Alignment(horizontal='center', vertical='center')

            # put border around non-opp cells; they will be formated later
            if col < 10:
                set_appearance(ws3, row + 1, col, 'fg_fill', 'FFFFFF')  # white

            col += 1

        # identify the duplicate files #todo: maybe do this when they come in? are they in order though?
        if hit[3] in file_seen:
            file_name_dups.append(hit[3])
            suite_dat.duplicate_file_name_hits.add(hit[3])
        else:
            file_seen.add(hit[3])

        row += 1

    # format_hit_data(suite_dat, hit_data, file_name_dups)
    format_hit_data(suite_dat, hit_data)


def get_test_case_name(hit_data):
    test_case_type = hit_data[1]
    file_name = hit_data[3]

    test_case_name = 'TEST_CASE_NAME_NOT_FOUND'

    if test_case_type == 'juliet':
        if suite_language == 'c':
            test_case_name = re.sub('[a-z]?\.\w+$', '', file_name)
        elif suite_language == 'cpp':
            test_case_name = re.sub('([a-z]?\.\w+$)|(_good.*)|(_bad.*)', '', file_name)
            # test_case_name = re.sub('([a-z]?\.\w+$)', '', file_name)
        else:
            print('ERROR_COULD_NOT_FIND_LANG')
    elif test_case_type == 'kdm':
        test_case_name = re.sub('[_a]?\.\w+$', '', file_name)
    else:
        print('ERROR_COULD_NOT_FIND_TYPE', test_case_type)

    return test_case_name


def format_hit_data(suite_dat, hit_data):
    row = 1
    start = 0
    group_size = 0
    next_group_idx = 0
    found_ops_in_group = []
    previous_file_name_and_line = []

    # color opportunities used(green), unused(red) or none(gray)
    for idx, hit in enumerate(hit_data):
        if idx == next_group_idx:
            found_ops_in_group = hit[9:13]
            test_case_name = get_test_case_name(hit)

            # a.
            group_size = suite_dat.suite_hit_data[test_case_name]
            # b.
            start = (idx + 2)
            # c.
            end = (idx + 2) + group_size - 1
            # d.
            next_group_idx = idx + group_size

            # todo: 5/4/17 skip merge if group size=1?
            # merge cells into test case groups
            for i in range(7, 14):
                ws3.merge_cells(start_row=start, start_column=i, end_row=end, end_column=i)

            # look thru all four possible opportunities
            for idx1, item in enumerate(found_ops_in_group):
                # opp not found = red
                for a in range(start, start + group_size):
                    set_appearance(ws3, a, idx1 + 10, 'fg_fill', 'FFC7CE')  # red
                # no opp = gray
                if not len(item):
                    for a in range(start, start + group_size):
                        set_appearance(ws3, a, idx1 + 10, 'fg_fill', 'D9D9D9')  # gray
        # opp found = green
        for idx1, item in enumerate(found_ops_in_group):
            if item in hit[5] and len(item) > 0:
                for a in range(start, start + group_size):
                    set_appearance(ws3, a, idx1 + 10, 'fg_fill', 'A9D08E')  # green

        # todo: 5/23/7 this needs optimized
        if hit[3] in suite_dat.duplicate_file_name_hits:
            if previous_file_name_and_line == list(hit[3:5]):
                #  yellow - repeat file name and line
                #  previous sorted file name and line are identical to this hit
                for hit_idx, item in enumerate(hit):
                    if hit_idx < 9:
                        # adjust current row
                        set_appearance(ws3, row + 1, hit_idx + 1, 'fg_fill', 'FFD966')  # yellow
                        # adjust previous row
                        set_appearance(ws3, row, hit_idx + 1, 'fg_fill', 'FFD966')  # yellow
            else:
                # blue - unique file name and line number
                for hit_idx, item in enumerate(hit):
                    if hit_idx < 9:
                        # adjust current row
                        set_appearance(ws3, row + 1, hit_idx + 1, 'fg_fill', 'BDD7EE')  # light blue

            previous_file_name_and_line = list(hit[3:5])

        row += 1


def import_xml_tags(suite_dat):
    row = 0

    ws = wb.get_sheet_by_name('XML Tags')
    row_count = ws.max_row
    col_count = ws.max_column

    tag_ids = [[0 for _ in range(col_count)] for _ in range(row_count)]

    for row_idx in ws.iter_rows():
        col = 0
        for cell in row_idx:
            tag_ids[row][col] = str(cell.value)
            col += 1
        row += 1

    setattr(suite_dat, 'tag_info', tag_ids)


def remove_dups(d):
    new_d = []
    for x in d:
        if x not in new_d:
            new_d.append(x)
    return new_d


def write_xml_data(suite_data_details):
    #########################################################################################################
    detail_sheet_titles = ['CWE', 'Type', 'T/F', 'TC', 'Hits', '%Hits', 'XML', 'TC Path', 'RAW Project File']
    #########################################################################################################

    row = 1
    attribute_list = ['cwe_id_padded', 'tc_type', 'true_false', 'tc_count', 'num_of_hits', 'percent_hits',
                      'new_xml_name', 'tc_path', 'scan_data_file']

    # write column headers
    for idx, title in enumerate(detail_sheet_titles):
        set_appearance(ws2, row, idx + 1, 'fg_fill', 'C9C9C9')
        ws2.cell(row=1, column=idx + 1).value = title
        ws2.cell(row=1, column=idx + 1).alignment = Alignment(horizontal='center')

    # write xml data
    for j, attrib in enumerate(attribute_list):
        for i, xml_data in enumerate(suite_data_details.xml_projects):
            # juliet or kdm
            tc_attrib = getattr(suite_data_details.xml_projects[i], attrib)

            # set colors for false
            if suite_data_details.xml_projects[i].true_false == 'FALSE':
                if suite_data_details.xml_projects[i].num_of_hits > 0:
                    # red
                    set_appearance(ws2, i + 2, 5, 'font_color', 'C00000')
                    set_appearance(ws2, i + 2, 6, 'font_color', 'C00000')
                    set_appearance(ws2, i + 2, j + 1, 'fg_fill', 'FFC7CE')
                else:
                    # green font, white fill
                    set_appearance(ws2, i + 2, 5, 'font_color', '548235')
                    set_appearance(ws2, i + 2, 6, 'font_color', '548235')
                    set_appearance(ws2, i + 2, j + 1, 'fg_fill', 'FFFFFF')

                # highlight juliet/false test cases as opp counts (vs. TC)
                if suite_data_details.xml_projects[i].tc_type == 'juliet':
                    set_appearance(ws2, i + 2, 4, 'font_color', '0000FF')
            else:
                # set colors for true
                if suite_data_details.xml_projects[i].num_of_hits > 0:
                    # green font, white fill
                    set_appearance(ws2, i + 2, 5, 'font_color', '548235')
                    set_appearance(ws2, i + 2, 6, 'font_color', '548235')
                    set_appearance(ws2, i + 2, j + 1, 'fg_fill', 'FFFFFF')
                else:
                    # red
                    set_appearance(ws2, i + 2, 5, 'font_color', 'C00000')
                    set_appearance(ws2, i + 2, 6, 'font_color', 'C00000')
                    set_appearance(ws2, i + 2, j + 1, 'fg_fill', 'FFC7CE')

            # write data to cells
            ws2.cell(row=i + 2, column=j + 1).value = tc_attrib

            # set fill color for col=1 and rows for hits=0
            if j == 0:
                # dark gray
                set_appearance(ws2, i + 2, j + 1, 'fg_fill', 'C6C6C6')

            # align columns
            if j == 1 or j == 2:
                ws2.cell(row=i + 2, column=j + 1).alignment = Alignment(horizontal='center')
            elif j == 6 or j == 7 or j == 8:
                ws2.cell(row=i + 2, column=j + 1).alignment = Alignment(horizontal='left')
            else:
                ws2.cell(row=i + 2, column=j + 1).alignment = Alignment(horizontal='right')


def write_summary_data(scan_data, ws):
    #########################################################################################################
    summary_sheet_titles = ['CWE', 'TC TRUE', 'TC FALSE', 'TP', 'FP', 'Precision', 'Recall']
    #########################################################################################################

    row = 1
    cwes = []

    # set all totals to zero
    suite_data.clear_totals()

    # write column headers
    for idx, title in enumerate(summary_sheet_titles):
        set_appearance(ws, row, idx + 1, 'fg_fill', 'E6B8B7')
        ws.cell(row=1, column=idx + 1).value = title
        ws.cell(row=1, column=idx + 1).alignment = Alignment(horizontal='center')

    for xml_project in scan_data.xml_projects:
        cwes.append(getattr(xml_project, 'cwe_id_padded'))

    unique_cwes = list(set(cwes))
    unique_cwes.sort()

    # collect data for each cwe and summarize
    for cwe in unique_cwes:
        suite_data.unique_cwes.append(cwe)

        tc_t = tc_f = tp = fp = 0
        row += 1

        # tally up the results from each xml project
        for xml_project in scan_data.xml_projects:
            if cwe == getattr(xml_project, 'cwe_id_padded'):
                if 'TRUE' == getattr(xml_project, 'true_false'):
                    tc_t += getattr(xml_project, 'tc_count')
                    tp += getattr(xml_project, 'num_of_hits')
                elif 'FALSE' == getattr(xml_project, 'true_false'):
                    tc_f += getattr(xml_project, 'tc_count')
                    fp += getattr(xml_project, 'num_of_hits')
                    # todo: add break or continue to speed this up?

        # for row1 in ws.iter_rows('A1:C2'):
        for row1 in ws.iter_rows():
            for cell in row1:
                if 1 < cell.col_idx < 9:
                    # highlight if no wid for this cwe
                    if all(x == 'None' for x in suite_data.acceptable_weakness_ids_full_list_dict[cwe]):
                        # light blue, no wid provided
                        set_appearance(ws, row, cell.col_idx, 'fg_fill', 'D6DCE4')
                    elif tp == 0:
                        # light gray, had at least one wid privided for the cwe, but no hits
                        set_appearance(ws, row, cell.col_idx, 'fg_fill', 'F2F2F2')
                    else:
                        set_appearance(ws, row, cell.col_idx, 'fg_fill', 'FFFFFF')

                    ws.cell(row=row, column=cell.col_idx).alignment = Alignment(horizontal='right')

        set_appearance(ws, row, 1, 'fg_fill', 'C6C6C6')

        # precision
        if tp + fp != 0:
            prec = tp / (tp + fp)
            ws.cell(row=row, column=6).value = round(prec, 2)
            ws.cell(row=row, column=6).number_format = '0.00'
        else:
            ws.cell(row=row, column=6).value = 'N/A'
            ws.cell(row=row, column=6).alignment = Alignment(horizontal='right')

        # recall
        recall = tp / tc_t

        # write cwe values to sheet
        ws.cell(row=row, column=1).value = cwe
        ws.cell(row=row, column=2).value = tc_t
        ws.cell(row=row, column=3).value = tc_f
        ws.cell(row=row, column=4).value = tp
        ws.cell(row=row, column=5).value = fp
        ws.cell(row=row, column=7).value = recall
        # update totals
        suite_data.suite_tc_count_true += tc_t
        suite_data.suite_tc_count_false += tc_f
        suite_data.suite_tp_count += tp
        suite_data.suite_fp_count += fp
        suite_data.suite_cwe_count = row - 1

        # apply format
        for col in range(2, 5):
            ws.cell(row=row, column=col).number_format = '#,##0'
        ws.cell(row=row, column=7).number_format = '0.00'

    # summary sheet
    if ws == ws1:
        ws.freeze_panes = ws['H2']
        write_score_and_message_to_summary(ws)
        # write_unweighted_averages(suite_data, ws)

    # score sheet
    if ws == ws5:
        # append columns to the right of current data
        # write_unweighted_averages(suite_data, ws)
        write_weighted_averages(ws)
        write_averages_to_summary_sheet()
        write_score_and_message_to_score_sheet(suite_data, ws)

    # write totals
    ws.cell(row=row + 1, column=2).value = suite_data.suite_tc_count_true
    ws.cell(row=row + 1, column=3).value = suite_data.suite_tc_count_false
    ws.cell(row=row + 1, column=4).value = suite_data.suite_tp_count
    ws.cell(row=row + 1, column=5).value = suite_data.suite_fp_count
    for col in range(2, 6):
        ws.cell(row=row + 1, column=col).number_format = '#,##0'
        set_appearance(ws, row + 1, col, 'font_color', '0000FF')  # blue


def write_averages_to_summary_sheet():
    # write helper col headers
    set_appearance(ws1, 1, 29, 'fg_fill', 'FFFFFF')
    set_appearance(ws1, 1, 30, 'fg_fill', 'FFFFFF')
    ws1.cell(row=1, column=29).value = 'Pavg= ' + '%0.2f' % suite_data.precision_average_unweighted
    ws1.cell(row=1, column=30).value = 'Ravg= ' + '%0.2f' % suite_data.recall_average_unweighted

    # write helper col p and r averages
    for row in ws1.iter_rows():
        if row[0].value is not None and row[0].row > 1:
            ws1.cell(row=row[0].row, column=29).value = suite_data.precision_average_unweighted
            set_appearance(ws1, row[0].row, 29, 'font_color', 'FFFFFF', False)
            ws1.cell(row=row[0].row, column=30).value = suite_data.recall_average_unweighted
            set_appearance(ws1, row[0].row, 30, 'font_color', 'FFFFFF', False)


def write_weighted_averages(ws):
    #########################################################################################
    score_sheet_titles_addendum = ['P-Wt.', 'P-Final', 'P-Avg', 'R-Wt.', 'R-Final', 'R-Avg.']
    #########################################################################################

    row = 1
    offset = 7

    # write column headers
    for idx, title in enumerate(score_sheet_titles_addendum):
        idx += offset
        set_appearance(ws, row, idx + 1, 'fg_fill', 'E6B8B7')
        ws.cell(row=1, column=idx + 1).value = title
        ws.cell(row=1, column=idx + 1).alignment = Alignment(horizontal='center')

    set_cwe_weightings(suite_data)

    for row_idx in ws.iter_rows():
        for cell in row_idx:
            if cell.col_idx == 1:
                if cell.row > 1:
                    # col 1 contains the cwe string
                    cwe = cell.value
                    weight = suite_data.weightings_per_cwe_dict[cwe]
                    # p-wt.
                    ws.cell(row=cell.row, column=offset + 1).value = weight
                    ws.cell(row=cell.row, column=offset + 1).number_format = '0.00'
                    ws.cell(row=cell.row, column=offset + 1).alignment = Alignment(horizontal='right')
                    set_appearance(ws, cell.row, offset + 1, 'fg_fill', 'DDEBF7')  # light blue
                    set_appearance(ws, cell.row, offset + 1, 'font_color', '833C0C')  # dark brown
                    # r-wt.
                    ws.cell(row=cell.row, column=offset + 4).value = weight
                    ws.cell(row=cell.row, column=offset + 4).number_format = '0.00'
                    ws.cell(row=cell.row, column=offset + 4).alignment = Alignment(horizontal='right')
                    set_appearance(ws, cell.row, offset + 4, 'fg_fill', 'EDEDED')  # light gray
                    set_appearance(ws, cell.row, offset + 4, 'font_color', '833C0C')  # dark brown

                    # p-final
                    if row_idx[5].value == 'N/A':
                        suite_data.precision_values_per_cwe_unweighted[cwe] = 'N/A'
                        ws.cell(row=cell.row, column=offset + 2).value = 'N/A'
                        set_appearance(ws, cell.row, offset + 2, 'font_color', '808080')  # med gray
                    else:
                        suite_data.precision_values_per_cwe_unweighted[cwe] = weight * row_idx[5].value
                        ws.cell(row=cell.row, column=offset + 2).value = suite_data.precision_values_per_cwe_unweighted[
                            cwe]
                        set_appearance(ws, cell.row, offset + 2, 'font_color', '0000FF')  # blue
                        ws.cell(row=cell.row, column=offset + 2).number_format = '0.00'
                    ws.cell(row=cell.row, column=offset + 2).alignment = Alignment(horizontal='right')
                    set_appearance(ws, cell.row, offset + 2, 'fg_fill', 'DDEBF7')  # light blue

                    # r-final
                    if row_idx[6].value == 0:
                        set_appearance(ws, cell.row, offset + 5, 'font_color', '808080')  # med gray
                    else:
                        set_appearance(ws, cell.row, offset + 5, 'font_color', 'C00000')  # dark red

                    suite_data.recall_values_per_cwe_unweighted[cwe] = weight * row_idx[6].value
                    ws.cell(row=cell.row, column=offset + 5).value = suite_data.recall_values_per_cwe_unweighted[cwe]
                    ws.cell(row=cell.row, column=offset + 5).number_format = '0.00'
                    ws.cell(row=cell.row, column=offset + 5).alignment = Alignment(horizontal='right')
                    set_appearance(ws, cell.row, offset + 5, 'fg_fill', 'EDEDED')  # light blue

                    # p-avg
                    if suite_data.precision_values_per_cwe_unweighted[cwe] != 'N/A':
                        suite_data.precision_accumulated_valid_count_unweighted += 1
                        suite_data.precision_accumulated_valid_values_unweighted += \
                            suite_data.precision_values_per_cwe_unweighted[cwe]
                        suite_data.precision_average_unweighted = \
                            suite_data.precision_accumulated_valid_values_unweighted \
                            / suite_data.precision_accumulated_valid_count_unweighted

                    # r-avg
                    suite_data.recall_accumulated_count_unweighted += 1
                    suite_data.recall_accumulated_values_unweighted += suite_data.recall_values_per_cwe_unweighted[cwe]
                    suite_data.recall_average_unweighted = \
                        suite_data.recall_accumulated_values_unweighted \
                        / suite_data.recall_accumulated_count_unweighted

    # p-avg display
    ws.merge_cells(start_row=2, start_column=offset + 3, end_row=ws.max_row, end_column=offset + 3)
    ws.cell(row=2, column=offset + 3).value = suite_data.precision_average_unweighted
    ws.cell(row=2, column=offset + 3).number_format = '0.00'
    ws.cell(row=2, column=offset + 3).alignment = Alignment(horizontal='center', vertical='center')
    set_appearance(ws, 2, offset + 3, 'font_color', '0000FF')  # blue
    set_appearance(ws, 2, offset + 3, 'fg_fill', 'DDEBF7')  # light blue
    set_appearance(ws, ws.max_row, offset + 3, 'fg_fill', 'DDEBF7')  # light blue
    # r-avg display
    ws.merge_cells(start_row=2, start_column=offset + 6, end_row=ws.max_row, end_column=offset + 6)
    ws.cell(row=2, column=offset + 6).value = suite_data.recall_average_unweighted
    ws.cell(row=2, column=offset + 6).number_format = '0.00'
    ws.cell(row=2, column=offset + 6).alignment = Alignment(horizontal='center', vertical='center')
    set_appearance(ws, 2, offset + 6, 'font_color', 'C00000')  # dark red
    set_appearance(ws, 2, offset + 6, 'fg_fill', 'EDEDED')  # light gray
    set_appearance(ws, ws.max_row, offset + 6, 'fg_fill', '808080')  # medium gray


def write_unweighted_averages(suite_data, ws):
    # todo: all calls to this function are currently disabled
    #########################################################################################
    score_sheet_titles_addendum = ['P-Wt.', 'P-Final', 'P-Avg', 'R-Wt.', 'R-Final', 'R-Avg.']
    #########################################################################################

    row = 1
    offset = 7

    # write column headers
    for idx, title in enumerate(score_sheet_titles_addendum):
        idx += offset
        set_appearance(ws, row, idx + 1, 'fg_fill', 'E6B8B7')
        ws.cell(row=1, column=idx + 1).value = title
        ws.cell(row=1, column=idx + 1).alignment = Alignment(horizontal='center')

    for row_idx in ws.iter_rows():
        for cell in row_idx:
            if cell.col_idx == 1:
                if cell.row > 1:
                    # col 1 contains the cwe string
                    cwe = cell.value
                    # weight = suite_data.weightings_per_cwe_dict[cwe]
                    weight = 1.00  # fixed for unweighted
                    # p-wt.
                    ws.cell(row=cell.row, column=offset + 1).value = weight
                    ws.cell(row=cell.row, column=offset + 1).number_format = '0.00'
                    ws.cell(row=cell.row, column=offset + 1).alignment = Alignment(horizontal='right')
                    set_appearance(ws, cell.row, offset + 1, 'fg_fill', 'DDEBF7')  # light blue
                    set_appearance(ws, cell.row, offset + 1, 'font_color', '833C0C')  # dark brown
                    # r-wt.
                    ws.cell(row=cell.row, column=offset + 4).value = weight
                    ws.cell(row=cell.row, column=offset + 4).number_format = '0.00'
                    ws.cell(row=cell.row, column=offset + 4).alignment = Alignment(horizontal='right')
                    set_appearance(ws, cell.row, offset + 4, 'fg_fill', 'EDEDED')  # light gray
                    set_appearance(ws, cell.row, offset + 4, 'font_color', '833C0C')  # dark brown
                    # p-final
                    if row_idx[5].value == 'N/A':
                        suite_data.precision_values_per_cwe_unweighted[cwe] = 'N/A'
                        ws.cell(row=cell.row, column=offset + 2).value = 'N/A'
                        set_appearance(ws, cell.row, offset + 2, 'font_color', '808080')  # med gray
                    else:
                        suite_data.precision_values_per_cwe_unweighted[cwe] = weight * row_idx[5].value
                        ws.cell(row=cell.row, column=offset + 2).value = suite_data.precision_values_per_cwe_unweighted[
                            cwe]
                        set_appearance(ws, cell.row, offset + 2, 'font_color', '0000FF')  # blue
                        ws.cell(row=cell.row, column=offset + 2).number_format = '0.00'
                    ws.cell(row=cell.row, column=offset + 2).alignment = Alignment(horizontal='right')
                    set_appearance(ws, cell.row, offset + 2, 'fg_fill', 'DDEBF7')  # light blue
                    # r-final
                    if row_idx[6].value == 0:
                        set_appearance(ws, cell.row, offset + 5, 'font_color', '808080')  # med gray
                    else:
                        set_appearance(ws, cell.row, offset + 5, 'font_color', 'C00000')  # dark red

                    suite_data.recall_values_per_cwe_unweighted[cwe] = weight * row_idx[6].value
                    ws.cell(row=cell.row, column=offset + 5).value = suite_data.recall_values_per_cwe_unweighted[cwe]
                    ws.cell(row=cell.row, column=offset + 5).number_format = '0.00'
                    ws.cell(row=cell.row, column=offset + 5).alignment = Alignment(horizontal='right')
                    set_appearance(ws, cell.row, offset + 5, 'fg_fill', 'EDEDED')  # light blue

                    # p-avg
                    if suite_data.precision_values_per_cwe_unweighted[cwe] != 'N/A':
                        suite_data.precision_accumulated_valid_count_unweighted += 1
                        suite_data.precision_accumulated_valid_values_unweighted += \
                            suite_data.precision_values_per_cwe_unweighted[cwe]
                        suite_data.precision_average_unweighted = suite_data.precision_accumulated_valid_values_unweighted \
                                                                  / suite_data.precision_accumulated_valid_count_unweighted

                    # r-avg
                    suite_data.recall_accumulated_count_unweighted += 1
                    suite_data.recall_accumulated_values_unweighted += suite_data.recall_values_per_cwe_unweighted[cwe]
                    suite_data.recall_average_unweighted = suite_data.recall_accumulated_values_unweighted \
                                                           / suite_data.recall_accumulated_count_unweighted

    # p-avg display
    ws.merge_cells(start_row=2, start_column=offset + 3, end_row=ws.max_row, end_column=offset + 3)
    ws.cell(row=2, column=offset + 3).value = suite_data.precision_average_unweighted
    ws.cell(row=2, column=offset + 3).number_format = '0.00'
    ws.cell(row=2, column=offset + 3).alignment = Alignment(horizontal='center', vertical='center')
    set_appearance(ws, 2, offset + 3, 'font_color', '0000FF')  # blue
    set_appearance(ws, 2, offset + 3, 'fg_fill', 'DDEBF7')  # light blue
    set_appearance(ws, ws.max_row, offset + 3, 'fg_fill', 'DDEBF7')  # light blue

    # r-avg display
    ws.merge_cells(start_row=2, start_column=offset + 6, end_row=ws.max_row, end_column=offset + 6)
    ws.cell(row=2, column=offset + 6).value = suite_data.recall_average_unweighted
    ws.cell(row=2, column=offset + 6).number_format = '0.00'
    ws.cell(row=2, column=offset + 6).alignment = Alignment(horizontal='center', vertical='center')
    set_appearance(ws, 2, offset + 6, 'font_color', 'C00000')  # dark red
    set_appearance(ws, 2, offset + 6, 'fg_fill', 'EDEDED')  # light gray
    set_appearance(ws, ws.max_row, offset + 6, 'fg_fill', '808080')  # medium gray


def set_cwe_weightings(suite_dat):
    # todo: all set to 1.0 for now
    for cwe in suite_dat.unique_cwes:
        suite_dat.weightings_per_cwe_dict[cwe] = 1.0
        # todo: 5/12/7 remove this, for testing only. assumption is that
        # todo: (cont) for each cwe, same weighting applies to both precisino and recall?


def write_score_and_message_to_summary(ws):
    # revision with git hash
    ws.cell(row=1, column=8).alignment = Alignment(horizontal='center', vertical='center')
    # ws.merge_cells(start_row=1, start_column=8, end_row=1, end_column=10)
    # todo: keep, this cell contains the hash/revision value of the code
    # ws.cell(row=1, column=8).value = ' \'score.exe\', v2.0.' + git_hash[:7]  # todo: keep short hash? or long?
    set_appearance(ws, 1, 8, 'font_color', '000000')  # black
    set_appearance(ws, 1, 8, 'fg_fill', 'F2F2F2')  # light gray
    # pass/fail notification
    ws.cell(row=1, column=9).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=1, column=9).value = suite_data.pass_fail
    set_appearance(ws, 1, 9, 'font_color', 'FFFFFF')  # white
    set_appearance(ws, 1, 9, 'fg_fill', '008000')  # green
    cell = ws['I1']
    cell.font = cell.font.copy(bold=True, italic=False)
    # manual review notification
    ws.cell(row=1, column=10).alignment = Alignment(horizontal='left', vertical='center')
    ws.merge_cells(start_row=1, start_column=10, end_row=1, end_column=28)
    ws.cell(row=1, column=10).value = suite_data.manual_review_recommendataion
    set_appearance(ws, 1, 10, 'font_color', '000000')  # black
    set_appearance(ws, 1, 10, 'fg_fill', 'F2F2F2')  # light gray
    cell = ws['J1']
    cell.font = cell.font.copy(bold=False, italic=True)


def write_score_and_message_to_score_sheet(suite_dat, ws):
    col_offset = 7

    # score label
    ws.cell(row=1, column=7 + col_offset).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=1, column=7 + col_offset).value = 'SCORE='
    set_appearance(ws, 1, 7 + col_offset, 'font_color', '000000')  # black
    set_appearance(ws, 1, 7 + col_offset, 'fg_fill', 'F2F2F2')  # light gray
    # score value
    ws.cell(row=1, column=8 + col_offset).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=1, column=8 + col_offset).value = '%0.2f' % (
        (suite_dat.precision_average_unweighted + suite_dat.recall_average_unweighted) / 2)
    set_appearance(ws, 1, 8 + col_offset, 'font_color', '000000')  # black
    set_appearance(ws, 1, 8 + col_offset, 'fg_fill', 'F2F2F2')  # light gray
    # threshold label
    ws.cell(row=1, column=9 + col_offset).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=1, column=9 + col_offset).value = 'THRESH='
    set_appearance(ws, 1, 9 + col_offset, 'font_color', '000000')  # black
    set_appearance(ws, 1, 9 + col_offset, 'fg_fill', 'FFE699')  # light yellow
    # threshold value
    ws.cell(row=1, column=10 + col_offset).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=1, column=10 + col_offset).value = suite_dat.overall_required_threshold_unweighted
    set_appearance(ws, 1, 10 + col_offset, 'font_color', '000000')  # black
    set_appearance(ws, 1, 10 + col_offset, 'fg_fill', 'F2F2F2')  # light gray
    # pass/fail notification
    ws.cell(row=1, column=11 + col_offset).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=1, column=11 + col_offset).value = suite_dat.pass_fail
    set_appearance(ws, 1, 11 + col_offset, 'font_color', 'FFFFFF')  # white
    set_appearance(ws, 1, 11 + col_offset, 'fg_fill', '008000')  # green


def set_appearance(ws_id, row_id, col_id, style_id, color_id, border=True):
    cell = ws_id.cell(row=row_id, column=col_id)

    if style_id == 'font_color':
        font_color = Font(color=color_id)
        cell.font = font_color

    if style_id == 'fg_fill':
        fill_color = PatternFill(fgColor=color_id, fill_type='solid')
        cell.fill = fill_color

    if border:
        # build thin border for all styles
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                             bottom=Side(style='thin'))
        cell.border = thin_border


def create_summary_charts():
    c2 = LineChart()
    # p-avg
    v2 = Reference(ws1, min_col=29, min_row=1, max_row=53)
    c2.add_data(v2, titles_from_data=True, from_rows=False)
    # r-avg
    v2 = Reference(ws1, min_col=30, min_row=1, max_row=53)
    c2.add_data(v2, titles_from_data=True, from_rows=False)
    c2.y_axis.scaling.min = 0
    c2.y_axis.scaling.max = 1
    # c2.legend = None
    c2.y_axis.crosses = 'max'

    # precision average
    s2 = c2.series[0]
    s2.graphicalProperties.line.solidFill = '4572A7'  # dark blue
    s2.graphicalProperties.line.dashStyle = 'dash'
    s2.graphicalProperties.line.width = 10000  # width in EMUs
    # recall average
    s2 = c2.series[1]
    s2.graphicalProperties.line.solidFill = 'C65911'  # dark orange
    s2.graphicalProperties.line.dashStyle = 'dash'
    s2.graphicalProperties.line.width = 10000  # width in EMUs

    # precision
    p_chart = BarChart(gapWidth=50)
    p_chart.type = 'col'
    # p_chart.style = 11 original
    p_chart.style = 5
    p_chart.y_axis.title = 'Precision & Recall'
    # p_chart.x_axis.title = 'CWE Number'

    recall_data = Reference(ws1, min_col=6, min_row=1, max_row=52, max_col=6)
    p_chart.add_data(recall_data, titles_from_data=True)
    recall_data = Reference(ws1, min_col=7, min_row=1, max_row=52, max_col=7)
    p_chart.add_data(recall_data, titles_from_data=True)

    # precision bars
    s5 = p_chart.series[0]
    s5.graphicalProperties.line.solidFill = 'FFFFFF'
    s5.graphicalProperties.solidFill = '4572A7'  # dark blue
    # recall bars
    s5 = p_chart.series[1]
    s5.graphicalProperties.line.solidFill = 'FFFFFF'
    s5.graphicalProperties.solidFill = '93A9CF'  # light blue

    cats = Reference(ws1, min_col=1, min_row=2, max_row=52)
    p_chart.set_categories(cats)
    p_chart.shape = 4
    p_chart.title = 'Protection Profile Scores (Precision & Recall) - Unweighted'
    p_chart.y_axis.scaling.min = 0
    p_chart.y_axis.scaling.max = 1
    p_chart.height = 13
    p_chart.width = 40

    p_chart += c2
    ws1.add_chart(p_chart, 'H2')

    tcc_true_bar_chart = BarChart(gapWidth=0)
    tcc_true_bar_chart.type = 'col'
    tcc_true_bar_chart.style = 5
    tcc_true_bar_chart.y_axis.title = 'Tese Case Counts (True)'

    tcc_true_data = Reference(ws1, min_col=2, min_row=1, max_row=52, max_col=2)
    tcc_true_bar_chart.add_data(tcc_true_data, titles_from_data=True)

    s33 = tcc_true_bar_chart.series[0]
    s33.graphicalProperties.line.solidFill = '000000'
    s33.graphicalProperties.line.width = 1000  # width in EMUs
    s33.graphicalProperties.solidFill = 'E6B8B7'  # light red

    cats = Reference(ws1, min_col=1, min_row=2, max_row=52)
    tcc_true_bar_chart.set_categories(cats)
    tcc_true_bar_chart.shape = 4
    tcc_true_bar_chart.title = 'Test Case Distribution'
    tcc_true_bar_chart.height = 10.4
    tcc_true_bar_chart.width = 40

    ws1.add_chart(tcc_true_bar_chart, 'H32')


def create_score_charts():
    p_offset = 3
    r_offset = 5

    p_r_average_line_chart = LineChart()
    # p-avg
    p_r_average_data = Reference(ws5, min_col=29, min_row=1, max_row=53)
    p_r_average_line_chart.add_data(p_r_average_data, titles_from_data=True, from_rows=False)
    # r-avg
    p_r_average_data = Reference(ws5, min_col=30, min_row=1, max_row=53)
    p_r_average_line_chart.add_data(p_r_average_data, titles_from_data=True, from_rows=False)
    # p and r average scaling
    p_r_average_line_chart.y_axis.scaling.min = 0
    p_r_average_line_chart.y_axis.scaling.max = 1
    p_r_average_line_chart.y_axis.crosses = 'max'
    # precision average
    p_r_average_series = p_r_average_line_chart.series[0]
    p_r_average_series.graphicalProperties.line.solidFill = '4572A7'  # dark blue
    p_r_average_series.graphicalProperties.line.dashStyle = 'dash'
    p_r_average_series.graphicalProperties.line.width = 10000  # width in EMUs
    # recall average
    p_r_average_series = p_r_average_line_chart.series[1]
    p_r_average_series.graphicalProperties.line.solidFill = 'C65911'  # dark orange
    p_r_average_series.graphicalProperties.line.dashStyle = 'dash'
    p_r_average_series.graphicalProperties.line.width = 10000  # width in EMUs

    # precision
    p_r_bar_chart = BarChart(gapWidth=50)
    p_r_bar_chart.type = 'col'
    p_r_bar_chart.style = 5
    p_r_bar_chart.y_axis.title = 'Precision & Recall'
    # p_r_bar_chart.x_axis.title = 'CWE Number'

    # precision
    p_r_data = Reference(ws5, min_col=6 + p_offset, min_row=1, max_row=52, max_col=6 + p_offset)
    p_r_bar_chart.add_data(p_r_data, titles_from_data=True)
    # recall
    p_r_data = Reference(ws5, min_col=7 + r_offset, min_row=1, max_row=52, max_col=7 + r_offset)
    p_r_bar_chart.add_data(p_r_data, titles_from_data=True)

    # precision bars
    s5 = p_r_bar_chart.series[0]
    s5.graphicalProperties.line.solidFill = '000000'
    s5.graphicalProperties.line.width = 1000  # width in EMUs
    # s5.graphicalProperties.solidFill = '4572A7'  # dark blue
    s5.graphicalProperties.solidFill = '548235'  # dark green
    # recall bars
    s5 = p_r_bar_chart.series[1]
    s5.graphicalProperties.line.solidFill = '000000'
    s5.graphicalProperties.line.width = 1000  # width in EMUs
    # s5.graphicalProperties.solidFill = '93A9CF'  # light blue
    s5.graphicalProperties.solidFill = 'A9D18E'  # light green

    cats = Reference(ws5, min_col=1, min_row=2, max_row=52)
    p_r_bar_chart.set_categories(cats)
    p_r_bar_chart.shape = 4
    p_r_bar_chart.title = 'Protection Profile Scores (Precision & Recall) - Weighted'
    p_r_bar_chart.y_axis.scaling.min = 0
    p_r_bar_chart.y_axis.scaling.max = 1
    p_r_bar_chart.height = 15
    p_r_bar_chart.width = 40

    # add charts
    p_r_bar_chart += p_r_average_line_chart
    ws5.add_chart(p_r_bar_chart, 'N2')

    tcc_true_bar_chart = BarChart(gapWidth=0)
    tcc_true_bar_chart.type = 'col'
    tcc_true_bar_chart.style = 5
    tcc_true_bar_chart.y_axis.title = 'Tese Case Counts (True)'

    tcc_true_data = Reference(ws5, min_col=2, min_row=1, max_row=52, max_col=2)
    tcc_true_bar_chart.add_data(tcc_true_data, titles_from_data=True)

    s33 = tcc_true_bar_chart.series[0]
    s33.graphicalProperties.line.solidFill = '000000'
    s33.graphicalProperties.line.width = 1000  # width in EMUs
    s33.graphicalProperties.solidFill = 'E6B8B7'  # light red

    cats = Reference(ws5, min_col=1, min_row=2, max_row=52)
    tcc_true_bar_chart.set_categories(cats)
    tcc_true_bar_chart.shape = 4
    tcc_true_bar_chart.title = 'Test Case Distribution'
    tcc_true_bar_chart.height = 11.3
    tcc_true_bar_chart.width = 40

    ws5.add_chart(tcc_true_bar_chart, 'N31')


def import_weakness_ids(suite_dat):
    # todo: consider consolidating this function with 'import_xml_tags'
    row = 0

    ws = wb.get_sheet_by_name('Weakness IDs')
    row_count = ws.max_row
    col_count = ws.max_column

    weakness_ids = [[0 for _ in range(col_count)] for _ in range(row_count)]

    # put all weakness ids into 'weakness_ids' list
    for row_idx in ws.iter_rows():
        col = 0
        for cell in row_idx:
            weakness_ids[row][col] = str(cell.value)
            col += 1

        # create a dictionary of wids per cwe
        if row > 0:
            cwe_id = 'CWE' + str(weakness_ids[row][0]).zfill(3)
            suite_data.acceptable_weakness_ids_full_list_dict[cwe_id] = weakness_ids[row][1:]

        row += 1
        # create a full list for the suite object
        setattr(suite_dat, 'acceptable_weakness_ids_full_list', weakness_ids)

    # add weakness ids to each xml object
    for i, xml_project in enumerate(suite_dat.xml_projects):
        cwe_num = getattr(xml_project, 'cwe_num')

        for weakness_id in weakness_ids:
            if weakness_id[0] == cwe_num:
                setattr(xml_project, 'acceptable_weakness_ids', weakness_id)
                break


def paint_wids_usage(used_wids, unused_wids):
    key = list(used_wids)[0]
    key = int(key[3:].lstrip('0'))

    # values = used wids
    values = list(used_wids.values())[0]
    unused_wid_values = list(unused_wids.values())[0]

    ws = wb.get_sheet_by_name('Weakness IDs')

    for row in ws.iter_rows():

        found_row = False
        for cell in row:
            # look for matching cwe row in sheet
            if key == cell.value:
                print('CWE_ID_ROW', cell.row, 'COL', cell.col_idx, cell.value)
                found_row = True
                continue

            if found_row:
                print('WID_ID', cell.row, 'COL', cell.col_idx, cell.value)
                # highlight all used wids green
                for value in values:
                    if value == cell.value:
                        # green = used
                        set_appearance(ws, cell.row, cell.col_idx, 'fg_fill', 'A9D08E')

                for value in unused_wid_values:
                    if value == cell.value:
                        # red
                        set_appearance(ws, cell.row, cell.col_idx, 'fg_fill', 'E6B8B7')


def get_unused_wids(scan_data, used_wids):
    acceptable_wids = {}

    cwe = list(used_wids)[0]
    cwe = cwe[3:].lstrip('0')

    for weak_id in scan_data.acceptable_weakness_ids_full_list:
        if weak_id[0] == cwe:
            del weak_id[0]
            acceptable_wids = {cwe: weak_id}

    used_values = list(used_wids.values())[0]
    acceptable_wids = list(acceptable_wids.values())[0]
    unused_wids = list(set(acceptable_wids) - set(used_values))

    if 'None' in unused_wids:
        unused_wids.remove('None')

    return {cwe: unused_wids}


def get_used_wids(scan_data):
    cwes = []

    for xml_project in scan_data.xml_projects:
        cwes.append(getattr(xml_project, 'cwe_id_padded'))

    unique_cwes = list(set(cwes))
    unique_cwes.sort()  # todo: necessary?

    for cwe in unique_cwes:
        used_wids_per_cwe = []

        # go thru each project looking for this cwe
        for xml_project in suite_data.xml_projects:
            # if the cwe for this project matches, get it's wids
            if cwe == getattr(xml_project, 'cwe_id_padded'):
                used_wids_per_cwe.extend(getattr(xml_project, 'used_wids'))
            else:
                continue

        unique_used_wids_per_cwe = list(set(used_wids_per_cwe))
        scan_data.used_wids_per_cwe.append([cwe, unique_used_wids_per_cwe])

        # todo: consider appending this dict like the list
        # scan_data.used_wids_per_cwe_dict = {cwe:unique_used_wids_per_cwe}
        used_wids = {cwe: unique_used_wids_per_cwe}
        unused_wids = get_unused_wids(scan_data, used_wids)

        paint_wids_usage(used_wids, unused_wids)


def githash(path):
    s = sha1()
    with open(path + '\\score.py', 'r') as f:
        while True:
            data1 = f.read().encode('utf-8')
            if not data1:
                break
    file_length = os.stat(path + '\\score.py').st_size
    s.update(('blob %u\0' % file_length).encode('utf-8'))
    s.update(data1)
    return s.hexdigest()


if __name__ == '__main__':
    py_common.print_with_timestamp('--- STARTED SCORING ---')
    parser = argparse.ArgumentParser(description='A script used to score all SCA tools.')
    # required
    parser.add_argument('language', help='The language being scanned (i.e. c, cpp or java)', type=str)
    parser.add_argument('suite', help='The suite number being scanned (i.e. 1 - 10)', type=int)
    # optional
    parser.add_argument('-n', dest='normalize', action='store_true', help='Enter \'\-n\' option for normalized score')

    args = parser.parse_args()
    suite_language = args.language
    suite_number = args.suite
    suite_path = os.getcwd()
    scaned_data_path = os.path.join(suite_path, 'scans')
    new_xml_path = os.path.join(suite_path, XML_OUTPUT_DIR)
    if args.normalize:
        normalize_juliet_false_scoring = True
    else:
        normalize_juliet_false_scoring = False

    # create scorecard from vendor input file
    time = strftime(
        'scorecard-fortify-' + suite_language + '_%m-%d-%Y_%H.%M.%S' + '_suite_' + str(suite_number).zfill(2))
    vendor_input = os.path.join(suite_path, 'vendor-input-' + TOOL_NAME + '-' + suite_language + '.xlsx')
    scorecard = os.path.join(suite_path, time) + '.xlsx'
    shutil.copyfile(vendor_input, scorecard)

    # get hash of score files for rev suffix
    # todo: this works but will also need to include all files in the build (not just score.py, but suite.py too)
    # git_hash = githash(suite_path)

    # add sheets and format
    wb = load_workbook(scorecard)
    ws1 = wb.create_sheet('Summary', 0)
    ws2 = wb.create_sheet('XML Data', 1)
    ws3 = wb.create_sheet('Hit Data', 2)
    ws4 = wb.create_sheet('Hit Analytics', 3)
    ws5 = wb.create_sheet('SCORE', 4)

    format_workbook()

    # instanciate a suite object and get suite data
    suite_data = Suite(scaned_data_path, new_xml_path, TOOL_NAME)

    # import tag data
    import_xml_tags(suite_data)
    # import weakness ids
    import_weakness_ids(suite_data)

    # score the xml projects
    score_xmls(suite_data)
    # get a summary of all used wids
    get_used_wids(suite_data)

    # write to sheets
    collect_hit_data(suite_data)
    write_xml_data(suite_data)

    # summary sheet
    write_summary_data(suite_data, ws1)
    # score sheet
    write_summary_data(suite_data, ws5)

    # chart for summary sheet
    create_summary_charts()
    create_score_charts()

    wb.active = 0
    wb.save(scorecard)

    py_common.print_with_timestamp('--- FINISHED SCORING ---')
